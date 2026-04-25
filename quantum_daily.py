import os, json, datetime, traceback

today      = datetime.date.today().strftime("%Y年%m月%d日")
today_file = datetime.date.today().strftime("%Y%m%d")

CAT_COLORS = {
    "熱伝導型":        ("FF6B35","FFEDE2"),
    "イオントラップ型": ("7B2D8B","F3E8F7"),
    "トポロジカル型":   ("00897B","E0F5F3"),
    "超伝導量子型":     ("1565C0","E3F0FB"),
    "光量子型":        ("F57C00","FFF3E0"),
    "国内エコシステム": ("2E7D32","E8F5E9"),
}
STATUS_COLORS = {
    "稼働中":"C6EFCE","開発中":"FFEB9C","研究段階":"FFDDC1",
    "実証実験":"E8DAEF","研究稼働中":"D4EDDA","商用稼働":"B7E4C7",
}

# ── STEP1: Web検索（失敗してもスキップ） ─────────────────
def step1_collect():
    print("\n[STEP1] Web検索開始")
    api_key = os.environ.get("ANTHROPIC_API_KEY","")
    if not api_key:
        print("  SKIP: ANTHROPIC_API_KEY 未設定")
        return []
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            tools=[{"type":"web_search_20250305","name":"web_search"}],
            messages=[{"role":"user","content":
                f"今日は{today}。最新の量子コンピューター性能情報を検索し"
                "新規発表/更新システムのみをJSON配列で返してください。"
                "なければ[]。キー:category,company,system_name,technology,"
                "qubit_count,qubit_count_note,two_qubit_gate_fidelity,"
                "coherence_time,key_features,announced_year,status。JSONのみ。"
            }],
        )
        for block in resp.content:
            if hasattr(block,"text"):
                text = block.text.strip()
                if "```" in text:
                    text = text.split("```")[1]
                    if text.startswith("json"): text = text[4:]
                try:
                    data = json.loads(text.strip())
                    result = data if isinstance(data,list) else []
                    print(f"  取得: {len(result)} 件")
                    return result
                except Exception:
                    print("  JSON解析失敗 → スキップ")
                    return []
    except Exception as e:
        print(f"  エラー（スキップ）: {e}")
    return []

# ── STEP2/3: Supabase 登録（失敗してもスキップ） ──────────
def step2_upsert(new_items):
    print("\n[STEP2/3] Supabase 登録・更新")
    if not new_items:
        print("  新規データなし → スキップ")
        return 0, 0
    try:
        from supabase import create_client
        sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])
        existing = {(r["system_name"],r["company"]):r["id"]
                    for r in sb.table("quantum_computers").select("id,system_name,company").execute().data}
        added = updated = 0
        fields = ["category","company","system_name","technology","qubit_count",
                  "qubit_count_note","two_qubit_gate_fidelity","coherence_time",
                  "key_features","announced_year","status"]
        for item in new_items:
            key = (item.get("system_name",""), item.get("company",""))
            payload = {k: item.get(k) for k in fields}
            if key in existing:
                sb.table("quantum_computers").update(payload).eq("id",existing[key]).execute()
                updated += 1
            else:
                sb.table("quantum_computers").insert(payload).execute()
                added += 1
        print(f"  追加:{added} 更新:{updated}")
        return added, updated
    except Exception as e:
        print(f"  エラー（スキップ）: {e}")
        traceback.print_exc()
        return 0, 0

# ── STEP4: Excel 生成（必ず実行） ─────────────────────────
def step4_excel():
    print("\n[STEP4] Excel 生成")
    try:
        from supabase import create_client
        sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])
        rows_raw = sb.table("quantum_computers").select("*")\
            .order("category").order("announced_year").order("company")\
            .execute().data
        print(f"  Supabaseから {len(rows_raw)} 件取得")
        ROWS = [(r.get("category",""),r.get("company",""),r.get("system_name",""),
                 r.get("technology",""),r.get("qubit_count"),r.get("qubit_count_note"),
                 r.get("two_qubit_gate_fidelity",""),r.get("coherence_time",""),
                 r.get("key_features",""),r.get("announced_year"),r.get("status",""))
                for r in rows_raw]
    except Exception as e:
        print(f"  Supabase取得エラー: {e}")
        traceback.print_exc()
        ROWS = []

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    def fill(h): return PatternFill("solid", fgColor=h)
    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="1F4E79")
    B_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    B_HD  = Border(left=thick,right=thick,top=thick,bottom=thick)

    wb = Workbook()
    ws = wb.active
    ws.title = "全件一覧"

    ws.merge_cells("A1:K1")
    c=ws["A1"]; c.value=f"最新量子コンピューター性能比較レポート ({today})"
    c.font=Font(name="Arial",size=15,bold=True,color="FFFFFF")
    c.fill=fill("1F4E79"); c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=28

    ws.merge_cells("A2:K2")
    c=ws["A2"]; c.value="Source: IBM, Google, Quantinuum, IonQ, Microsoft, D-Wave, RIKEN, AIST (auto-updated daily)"
    c.font=Font(name="Arial",size=8,italic=True,color="595959")
    c.fill=fill("EBF3FB"); c.alignment=Alignment(horizontal="center")
    ws.row_dimensions[2].height=14; ws.row_dimensions[3].height=5

    HDRS = ["カテゴリ","企業名","システム名","技術方式","量子ビット数","備考",
            "2QB忠実度","コヒーレンス時間","主な特徴","発表年","ステータス"]
    WIDTHS = [18,24,22,18,12,14,14,16,44,10,12]
    for ci,(h,w) in enumerate(zip(HDRS,WIDTHS),1):
        c=ws.cell(row=4,column=ci,value=h)
        c.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); c.fill=fill("2E75B6")
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=B_HD
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[4].height=22

    prev=None
    for ri,row in enumerate(ROWS,start=5):
        cat,stat=row[0],row[10]
        hx,bg=CAT_COLORS.get(cat,("555555","F2F2F2")); sb_=STATUS_COLORS.get(stat,"FFFFFF")
        vals=list(row); vals[4]=row[4] if row[4] else "—"; vals[5]=row[5] if row[5] else "—"
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=ri,column=ci,value=val)
            if ci==1: c.font=Font(name="Arial",size=9,bold=True,color=hx); c.fill=fill(bg)
            elif ci==11: c.font=Font(name="Arial",size=9,bold=True); c.fill=fill(sb_)
            elif ci==5 and isinstance(val,int):
                c.font=Font(name="Arial",size=9,bold=True); c.fill=fill("FFFFFF"); c.number_format="#,##0"
            else: c.font=Font(name="Arial",size=9); c.fill=fill("FFFFFF")
            c.alignment=Alignment(horizontal="left" if ci==9 else "center",vertical="center",wrap_text=(ci==9))
            c.border=Border(left=thick if ci==1 else thin, right=thick if ci==11 else thin,
                            top=Side(style="medium" if cat!=prev else "thin",
                                     color="1F4E79" if cat!=prev else "BFBFBF"),bottom=thin)
        ws.row_dimensions[ri].height=26; prev=cat

    # ランキングシート
    ws2=wb.create_sheet("量子ビット数ランキング")
    ws2.merge_cells("A1:D1"); c=ws2["A1"]
    c.value=f"量子ビット数ランキング ({today})"
    c.font=Font(name="Arial",size=13,bold=True,color="FFFFFF"); c.fill=fill("1F4E79")
    c.alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=24
    for ci,(h,w) in enumerate(zip(["順位","システム名（企業）","カテゴリ","量子ビット数"],[8,36,18,16]),1):
        c=ws2.cell(row=3,column=ci,value=h)
        c.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); c.fill=fill("2E75B6")
        c.alignment=Alignment(horizontal="center"); c.border=B_ALL
        ws2.column_dimensions[get_column_letter(ci)].width=w
    rd=sorted([(r[2],r[1],r[0],r[4]) for r in ROWS if isinstance(r[4],int)],key=lambda x:-x[3])
    for ri,(sn,co,cat,qb) in enumerate(rd,1):
        hx,bg=CAT_COLORS.get(cat,("555555","F2F2F2")); rn=ri+3
        for ci,val in enumerate([ri,f"{sn} ({co})",cat,qb],1):
            c=ws2.cell(row=rn,column=ci,value=val)
            if ci==3: c.font=Font(name="Arial",size=9,bold=True,color=hx); c.fill=fill(bg)
            elif ci==4: c.font=Font(name="Arial",size=9,bold=True); c.number_format="#,##0"; c.fill=fill("FFFFFF")
            else: c.font=Font(name="Arial",size=9); c.fill=fill("F2F2F2" if ci==1 else "FFFFFF")
            c.alignment=Alignment(horizontal="left" if ci==2 else "center"); c.border=B_ALL
        ws2.row_dimensions[rn].height=20
    if rd:
        bar=BarChart(); bar.type="bar"; bar.style=10; bar.height=16; bar.width=24
        bar.add_data(Reference(ws2,min_col=4,min_row=3,max_row=3+len(rd)),titles_from_data=True)
        bar.set_categories(Reference(ws2,min_col=2,min_row=4,max_row=3+len(rd)))
        bar.series[0].graphicalProperties.solidFill="2E75B6"; ws2.add_chart(bar,"F3")

    out=f"quantum_computers_{today_file}.xlsx"
    wb.save(out); print(f"  保存完了: {out} ({len(ROWS)}件)")
    return out

# ── STEP5: Slack 投稿（失敗してもスキップ） ───────────────
def step5_slack(added, updated):
    print("\n[STEP5] Slack 投稿")
    try:
        from slack_sdk import WebClient
        info = f"本日: 新規{added}件追加 / {updated}件更新" if (added or updated) else "本日: 新規追加・更新なし"
        msg = (
            f"🔬 *量子コンピューター 性能ランキング* — {today}\n{info}\n\n"
            "━━━━━━━━━━━━━━━━━━━\n🏆 カテゴリ別 性能トップシステム\n━━━━━━━━━━━━━━━━━━━\n"
            "1️⃣ *トポロジカル型* — Microsoft「Majorana 1」\n   └ 論理QB忠実度：理論上 >99.99%（2025年世界初実証）\n"
            "2️⃣ *イオントラップ型* — Quantinuum「H2-1」\n   └ 2QB忠実度：>99.9% ／ コヒーレンス時間：約1時間\n"
            "3️⃣ *超伝導量子型* — IBM「Heron r2」\n   └ 156量子ビット ／ 2QB忠実度：~99.9%\n"
            "4️⃣ *光量子型* — Xanadu「Borealis」\n   └ 216スクイーズドモード ／ 量子優位性を実証済\n"
            "5️⃣ *熱伝導型* — D-Wave「Advantage2」\n   └ 7,000量子ビット ／ 組合せ最適化で商用トップ実績\n"
            "6️⃣ *国内エコシステム* — 富士通＋理研「256QB機」\n   └ 国産最大規模 ／ 産総研（QuAI）も超伝導QB研究を推進\n"
            "━━━━━━━━━━━━━━━━━━━\n📊 詳細: Supabase `quantum_computers` テーブル"
        )
        token = os.environ.get("SLACK_BOT_TOKEN","")
        ch    = os.environ.get("SLACK_CHANNEL_ID","CA4N4ETA7")
        WebClient(token=token).chat_postMessage(channel=ch, text=msg, mrkdwn=True)
        print("  投稿完了")
    except Exception as e:
        print(f"  エラー（スキップ）: {e}")
        traceback.print_exc()

# ── メイン ─────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n{'='*50}\n量子コンピューター 毎日自動更新 — {today}\n{'='*50}")
    try:
        new_items      = step1_collect()
        added, updated = step2_upsert(new_items)
        excel_path     = step4_excel()          # ← 必ず実行
        step5_slack(added, updated)
        print(f"\n✅ 完了: 追加:{added} 更新:{updated} ファイル:{excel_path}")
    except Exception as e:
        print(f"\n❌ 予期しないエラー: {e}")
        traceback.print_exc()
        raise
